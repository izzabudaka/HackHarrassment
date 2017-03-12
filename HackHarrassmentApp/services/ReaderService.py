import os
from openpyxl import load_workbook
import xml.etree.ElementTree


class ReaderService:
    def __init__(self):
        self.conversation_labels = {}
        self.conversation_text = {}
        self.RESOURCE_DIR = os.path.dirname(os.path.abspath(__file__)) + "/../../resources/"
        self.LABELS_DIR = self.RESOURCE_DIR + "labels/"
        self.CONVERSATIONS_DIR = self.RESOURCE_DIR + "conversations/"

    def read_bad_words(self):
        return open(self.RESOURCE_DIR + "banned_words.txt").read().split("\n")

    def read_labels(self):
        label_files = os.listdir(self.LABELS_DIR)
        for file in label_files:
            if not file.startswith("."):
                self.read_label_file(self.LABELS_DIR + file)

    def read_other_data_file(self):
        formed = xml.etree.ElementTree.parse(self.RESOURCE_DIR + "XMLMergedFile.xml").getroot().findall('FORMSPRINGID')
        current = 0
        for form in formed:
            post = form.find('POST')
            text = post.find('TEXT').text
            labels = post.findall('LABELDATA')
            total = 0
            for label in labels:
                total += 1 if label.find('ANSWER').text == 'Yes' else 0
            if total > 2:
                self.conversation_labels[current] = 'Y'
            else:
                self.conversation_labels[current] = 'N'
            self.conversation_text[current] = text
            current += 1
        pass

    def read_data_files(self):
        convo_group_files = os.listdir(self.CONVERSATIONS_DIR)
        for file in convo_group_files:
            if not file.startswith("."):
                current_group = self.CONVERSATIONS_DIR + file + "/"
                convo_files = os.listdir(current_group)
                for convo_file in convo_files:
                    if not convo_file.startswith("."):
                        self.read_data_file(convo_file, current_group + convo_file)
        pass

    def read_label_file(self, filename):
        workbook = load_workbook(filename)
        worksheet = workbook.get_sheet_names()[0]
        worksheet = workbook.get_sheet_by_name(worksheet)
        idx = 0
        for row in worksheet.iter_rows():
            if idx > 2:
                values = [cell._value for cell in row[0:2]]
                if values[0] is not None:
                    self.conversation_labels[float(values[0])] = values[1]
            idx += 1

    def get_harrassment_convos(self):
        all_keys = self.conversation_labels.keys()
        return [key for key in all_keys if key in self.conversation_text and self.conversation_labels[key] == 'Y']

    def get_all_convos(self):
        all_keys = self.conversation_labels.keys()
        return [key for key in all_keys if key in self.conversation_text]

    def read_data_file(self, filename, convo_file_dir):
        try:
            posts = xml.etree.ElementTree.parse(convo_file_dir).getroot().findall('post')
            all_text = ""
            for post in posts:
                text = post.find('body').text
                if text is not None:
                    all_text += text.lower()
            self.conversation_text[float(".".join(filename.split(".")[0:2]))] = all_text
        except:
            print("Couldn't parse ", filename)
